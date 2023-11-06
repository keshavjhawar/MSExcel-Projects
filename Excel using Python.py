port openpyxl
from openpyxl import Workbook,load_workbook

wb = Workbook()
wb.save("Automation.xlsx")
wb1 = load_workbook('/Users/macos/PycharmProjects/pythonProject2/Automation.xlsx')
ws = wb1.active
print(ws)
wb1_sheet = wb1['Sheet']
wb1_sheet.title = 'Code with Keshav'
wb1_sheet = wb1['Code with Keshav']
wb1_sheet.sheet_properties.tabColor = 'FF0000'
wb1.save('/Users/macos/PycharmProjects/pythonProject2/Automation.xlsx')

import xlsxwriter
wb2 = xlsxwriter.Workbook('/Users/macos/PycharmProjects/Python+Excel/Python_Excel.xlsx')
wb2.close()

import openpyxl
wb3 = openpyxl.load_workbook('/Users/macos/PycharmProjects/Python+Excel/Python_Excel.xlsx')
ws = wb3.active
ws["A1"].value = "Keshav"

from openpyxl.styles import Alignment
ws.merge_cells("A1:B2")
ws.unmerge_cells("A1:B2")

ws.row_dimensions[1].height = 50
ws.column_dimensions['B'].width = 8

wb3.save('/Users/macos/PycharmProjects/Python+Excel/Python_Excel.xlsx')
from pickle import TRUE

import pandas as pd
import numpy as np


data = np.array([[1,2],[3,4],[5,6]])

df = pd.DataFrame(data, index=['row1','row2','row3'], columns=['col1','col2'])

print(df)


data = [[1,2],[3,4],[5,6]]

df = pd.DataFrame(data, index=['row1','row2','row3'], columns=['col1','col2'])

print(df)

states = ['Rajasthan','Maharashtra','Gujarat']
population = [68548437,126189673,60439692]

dict_states = {'States': states, 'Population': population}

df_pop = pd.DataFrame(dict_states)

print(df_pop)
import IPython
from IPython.display import display
from matplotlib import pyplot as plt

pd.set_option('display.max_columns', 20)


students_data = pd.read_csv('/Users/macos/PycharmProjects/Python+Excel/StudentsPerformance.csv')
display(students_data.head(7))


import numpy as np

lang_score = np.arange(0,1000)
students_data['language score'] = lang_score
display(students_data)

lang_score = np.random.randint(1,101,size=1000)
students_data['language score'] = lang_score
display(students_data)

total_maths = students_data['math score'].sum()
print(total_maths)

total_score = students_data['math score'] + students_data['reading score'] + students_data['writing score']
# print(total_score)
students_data['Total Score'] = total_score
avg_score = total_score/3
students_data['Avg Score'] = avg_score
# display(students_data)

total_avg_score = students_data['Avg Score'].mean()
display(students_data)

print(students_data['lunch'].value_counts(normalize=TRUE))
Sorted = students_data.sort_values(['parental level of education'], inplace = True,ascending=True, key= lambda col:col.str.lower())
# print(Sorted)

students_data.to_csv('/Users/macos/PycharmProjects/Python+Excel/StudentsPerformance.csv')
total_fail = 0
total_pass = 0
for score in students_data['Avg Score']:
    if score >= 36:
        total_pass +=1

    else:
        total_fail += 1
print('Total passed student =',total_pass)
print('Total failed student =',total_fail)

df_gdp = pd.read_csv('/Users/macos/PycharmProjects/Python+Excel/gdp.csv',encoding='latin-1')
pivot1 = df_gdp.pivot_table(index = ['year'],columns=['country'],values = 'gdppc')

for year in pivot1['India'].index:
  if pivot1.loc[int(year),'India'] > 600:
      print(year,':',pivot1.loc[int(year),'India'])

import matplotlib
pivot1['India'].plot(kind='bar',stacked=True)
pivot1['United Kingdom'].plot(kind='bar',stacked=True)
pivot1['United States'].plot(kind='bar',stacked=True)
plt.show()

pivot1_2010 = pivot1[pivot1.index.isin([1980,2010])]
pivot1_2010 = pivot1_2010.T

pivot1_2010.head(4).plot(kind='bar')
plt.show()
